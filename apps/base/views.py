
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from django.http import HttpResponse

def decode_xls_file(xls_file):
    try:
        workbook = openpyxl.load_workbook(xls_file)
        sheet = workbook.active
        
        header = [cell.value for cell in sheet[1]]
        
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(list(row))
        
        return header, rows, None
    except Exception as e:
        return None, None, f"Error al procesar el archivo Excel: {e}"

def validate_columns(header, expected_columns):
    if header != expected_columns:
        return False, f"El archivo Excel debe tener las columnas {', '.join(expected_columns)} en el orden correcto."
    return True, None

def process_xls_rows(rows, expected_columns_length):
    for row in rows:
        if len(row) != expected_columns_length:
            return None, f"Fila con número incorrecto de columnas: {row}"
    return rows, None


def apply_cell_format(cell, header=False):
    header_fill = PatternFill(start_color='FA603B', end_color='FA603B', fill_type='solid')
    header_font = Font(color='FFFFFF')  
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if header:
        cell.fill = header_fill
        cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")

def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

def export_data(request, title, headers, data, filename):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title

    worksheet.append(headers)
    for cell in worksheet[1]:
        apply_cell_format(cell, header=True)

    for row_idx, row in enumerate(data, start=2):
        worksheet.append(row)
        for cell in worksheet[row_idx]:
            apply_cell_format(cell)

    adjust_column_width(worksheet)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    workbook.save(response)
    
    return response
