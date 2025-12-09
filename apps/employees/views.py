from django.contrib import messages
from django.http import HttpResponse
from django.db import transaction
from datetime import datetime
import pandas as pd
import unicodedata
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from apps.employees.models import Employee

def normalize_column(col):
    col = col.lower().strip()
    col = ''.join(c for c in unicodedata.normalize('NFKD', col) if not unicodedata.combining(c))
    return col

def export_employee(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Empleados'

    # Header styling
    header_fill = PatternFill(start_color='FA603B', end_color='FA603B', fill_type='solid')
    header_font = Font(color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define headers
    headers = ['Tipo de documento', 'Número de documento', 'Nombre completo', 'Correo', 'Celular', 'Género', 'Fecha de ingreso']
    worksheet.append(headers)

    # Style header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fetch employee data and populate rows
    employees = Employee.objects.all()
    for row_idx, employee in enumerate(employees, start=2):
        row = [
            employee.document_type,
            employee.document_number,
            employee.full_name,
            employee.email,
            employee.cellphone,
            employee.gender,
            employee.date_entry.strftime('%d/%m/%Y') if employee.date_entry else ""
        ]
        worksheet.append(row)

        # Style each cell in the row
        for cell in worksheet[row_idx]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust column widths
    for col in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        column_letter = col[0].column_letter
        worksheet.column_dimensions[column_letter].width = max_length + 2

    # Generate HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Empleados.xlsx'
    workbook.save(response)

    return response
