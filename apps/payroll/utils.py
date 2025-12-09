
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from apps.payroll.models import PayrollEmployee, PayrrollEmployeeDetail
from django.http import HttpResponse
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from django.http import HttpResponse
from collections import defaultdict
from reportlab.pdfbase.pdfmetrics import stringWidth

def export_payroll_excel_response(payroll):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    # Estilos
    bold_center = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Título
    title = f"Nómina-{payroll.month}-{payroll.year} | Año: {payroll.year} - Mes: {payroll.month}"
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    # Subtítulo con total
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Total: {float(payroll.total_amount):,.2f}"
    ws['A2'].alignment = center_align
    ws['A2'].font = Font(italic=True)

    # Obtener empleados
    employees = PayrollEmployee.objects.filter(payroll=payroll).prefetch_related(
        'employee__position',
        'employee__area',
        'Nomina_Empleado__concept'
    )

    # Obtener headers de conceptos únicos
    concept_set = set()
    for emp in employees:
        for detail in emp.Nomina_Empleado.all():
            if detail.concept:
                concept_set.add(detail.concept.description or detail.concept.name)
    concept_headers = sorted(filter(None, concept_set))

    # Headers
    headers = ['#', 'DNI - Empleado', 'Área', 'Cargo', 'Fecha de ingreso'] + concept_headers
    ws.append(headers)

    # Pintar headers con estilos
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = column_title
        cell.font = bold_center
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Fila de empleados
    concept_totals = {concept: 0 for concept in concept_headers}
    for row_idx, emp in enumerate(employees, start=4):
        row = [
            row_idx - 3,
            f"{emp.employee.document_number} - {emp.employee.full_name}",
            emp.employee.area.name if emp.employee and emp.employee.area else '',
            emp.employee.position.name if emp.employee and emp.employee.position else '',
            emp.employee.date_entry.strftime("%Y-%m-%d") if emp.employee and emp.employee.date_entry else ''
        ]

        concept_map = {
            detail.concept.description or detail.concept.name: float(detail.calculated_amount)
            for detail in emp.Nomina_Empleado.all() if detail.concept
        }

        for concept in concept_headers:
            value = concept_map.get(concept, 0)
            row.append(value)
            concept_totals[concept] += value

        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = value
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            cell.alignment = center_align
            cell.border = thin_border

    total_row_idx = len(employees) + 4
    for i in range(1, 6):  
        cell = ws.cell(row=total_row_idx, column=i)
        cell.value = '' if i != 5 else 'Totales:'
        cell.font = bold_center if i == 5 else None
        cell.alignment = center_align
        cell.border = thin_border

    for idx, concept in enumerate(concept_headers, start=6):
        cell = ws.cell(row=total_row_idx, column=idx)
        cell.value = concept_totals[concept]
        cell.font = bold_center
        cell.number_format = '#,##0.00'
        cell.alignment = center_align
        cell.border = thin_border

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Nomina_{payroll.month}_{payroll.year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def draw_concepto_y_monto(p, x, y, concepto, monto, col_width, font="Helvetica", size=8):
    max_width = col_width - 1.8 * cm  # espacio para el monto
    p.setFont(font, size)
    while stringWidth(concepto, font, size) > max_width and len(concepto) > 3:
        concepto = concepto[:-1]
    if stringWidth(concepto, font, size) > max_width:
        concepto = concepto[:-3] + "..."
    p.drawString(x, y, concepto)
    p.drawRightString(x + col_width - 0.2 * cm, y, f"{monto:.2f}")

def export_payroll_boleta_response(payroll_employee):
    empleado = payroll_employee.employee
    nombre = empleado.full_name or '---'
    dni = empleado.document_number or '---'
    cargo = getattr(empleado.position, 'name', '---') if empleado.position else '---'
    fecha_ingreso = empleado.date_entry.strftime('%d/%m/%Y') if empleado.date_entry else '---'
    codigo_afp = empleado.code_afp or '---'
    type_pension = empleado.type_pension or '---'
    neto = payroll_employee.net_total
    dias_trabajados = payroll_employee.days_worked or 0
    horas_trabajadas = payroll_employee.total_hours or 0
    dias_vacaciones = 0
    dias_descanso = 0
    faltas = payroll_employee.faults
    horas_extras = payroll_employee.overtime
    mes_nombre = payroll_employee.payroll.month
    year = payroll_employee.payroll.year

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setTitle(f"Boleta de Pago - {nombre} - {mes_nombre} {year}")
    width, height = A4

    # Encabezado
    p.setFont("Helvetica-Bold", 22)
    p.drawString(2 * cm, height - 2.5 * cm, "pattern")
    p.setFont("Helvetica-Bold", 10)
    p.drawRightString(width - 2 * cm, height - 2.5 * cm, "PATTERN GROUP S.A.C")
    p.drawRightString(width - 2 * cm, height - 3.0 * cm, "20604442533")
    p.drawRightString(width - 2 * cm, height - 3.5 * cm, "BOLETA DE PAGO")

    # Datos empleado
    p.setFont("Helvetica", 8)
    y = height - 4.2 * cm
    p.drawString(2 * cm, y, f"CÓDIGO: CT-{dni}")
    p.drawString(2 * cm, y - 0.5 * cm, f"DNI: {dni}")
    p.drawString(2 * cm, y - 1.0 * cm, f"APELLIDOS Y NOMBRES: {nombre}")
    p.drawString(2 * cm, y - 1.5 * cm, f"CARGO: {cargo}")
    p.drawString(11 * cm, y, f"FECHA INGRESO: {fecha_ingreso}")
    p.drawString(11 * cm, y - 0.5 * cm, f"CÓDIGO DE AFP: {codigo_afp}")
    p.drawString(11 * cm, y - 1.0 * cm, f"TIPO DE PENSIÓN: {type_pension}")

    # Línea separadora
    y -= 2.2 * cm
    p.line(2 * cm, y, width - 2 * cm, y)

    # Resumen labores
    y -= 1.0 * cm
    p.setFont("Helvetica-Bold", 9)
    p.drawString(2 * cm, y, f"RESUMEN DE LABORES DEL MES DE {mes_nombre.upper()} {year}")
    y -= 0.8 * cm
    p.setFont("Helvetica", 8)
    p.drawString(2 * cm, y, f"DÍAS TRABAJADOS: {dias_trabajados}")
    p.drawString(2 * cm, y - 0.5 * cm, f"HORAS TRABAJADAS: {horas_trabajadas}")
    p.drawString(2 * cm, y - 1.0 * cm, f"DÍAS DE VACACIONES: {dias_vacaciones}")
    p.drawString(11 * cm, y, f"DÍAS DE DESCANSO: {dias_descanso}")
    p.drawString(11 * cm, y - 0.5 * cm, f"FALTAS: {faltas}")
    p.drawString(11 * cm, y - 1.0 * cm, f"HORAS EXTRAS: {horas_extras}")

    # Línea separadora
    y -= 2.2 * cm

    # Conceptos nómina
    detalles = payroll_employee.Nomina_Empleado.all()
    datos = defaultdict(list)
    for d in detalles:
        datos[d.concept.type.lower()].append((d.concept.name, d.calculated_amount))

    ingresos = datos.get('base', [])
    descuentos = datos.get('descuento', [])
    aportes = datos.get('auxiliar', [])

    col_x = [2 * cm, 9 * cm, 14.5 * cm]
    col_w = [6.5 * cm, 5 * cm, 4.5 * cm]
    headers = ["INGRESOS DEL TRABAJADOR", "DESCUENTOS AL TRABAJADOR", "APORTES DEL EMPLEADOR"]

    y_tabla = y - 1 * cm
    p.setFont("Helvetica-Bold", 8)
    for i, header in enumerate(headers):
        p.drawString(col_x[i], y_tabla, header)

    y_tabla -= 0.5 * cm
    p.setFont("Helvetica", 8)
    max_filas = max(len(ingresos), len(descuentos), len(aportes))

    for i in range(max_filas):
        if i < len(ingresos):
            draw_concepto_y_monto(p, col_x[0], y_tabla, ingresos[i][0], ingresos[i][1], col_w[0])
        if i < len(descuentos):
            draw_concepto_y_monto(p, col_x[1], y_tabla, descuentos[i][0], descuentos[i][1], col_w[1])
        if i < len(aportes):
            draw_concepto_y_monto(p, col_x[2], y_tabla, aportes[i][0], aportes[i][1], col_w[2])
        y_tabla -= 0.4 * cm

    # Totales
    p.setFont("Helvetica-Bold", 8)
    p.drawRightString(col_x[0] + col_w[0] - 0.2 * cm, y_tabla, f"{sum(x[1] for x in ingresos):.2f}")
    p.drawRightString(col_x[1] + col_w[1] - 0.2 * cm, y_tabla, f"{sum(x[1] for x in descuentos):.2f}")
    p.drawRightString(col_x[2] + col_w[2] - 0.2 * cm, y_tabla, f"{sum(x[1] for x in aportes):.2f}")

    # Línea separadora
    y_tabla -= 0.6 * cm
    p.line(2 * cm, y_tabla, width - 2 * cm, y_tabla)

    # Neto
    y_tabla -= 1 * cm
    p.setFont("Helvetica-Bold", 10)
    p.drawRightString(width - 2 * cm, y_tabla, f"NETO A PAGAR: S/ {neto:.2f}")

    # Firma gerente con imagen
    y_tabla -= 4 * cm
    firma_path = "static/admin/img/FIRMA.jpeg"
    firma_w = 6 * cm
    firma_h = 3.2 * cm
    firma_x = 12 * cm
    firma_y = y_tabla
    try:
        p.drawImage(firma_path, firma_x, firma_y, width=firma_w, height=firma_h, mask='auto')
    except:
        pass

    # Firmas
    p.line(2 * cm, y_tabla, 8 * cm, y_tabla)
    p.drawString(2 * cm, y_tabla - 0.4 * cm, nombre)
    p.drawString(2 * cm, y_tabla - 0.8 * cm, "RECIBÍ CONFORME")

    p.line(firma_x, y_tabla, width - 2 * cm, y_tabla)
    p.drawString(firma_x, y_tabla - 0.4 * cm, "Juan Carlos Gonzales Quispe")
    p.drawString(firma_x, y_tabla - 0.8 * cm, "Gerente General")

    # Finalizar
    p.showPage()
    p.save()
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Boleta_{dni}_{mes_nombre}_{year}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response
